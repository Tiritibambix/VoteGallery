import os
import uuid
import sqlite3
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, request, jsonify, make_response, send_from_directory, send_file
import base64
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__, static_folder='static', template_folder='templates')

# Créer le répertoire de données s'il n'existe pas
DATA_DIR = os.path.join(os.getcwd(), 'data')
os.makedirs(DATA_DIR, exist_ok=True)

DB_PATH = os.path.join(DATA_DIR, 'votes.db')
PHOTOS_DIR = 'photos'
ADMIN_PASSWORD = os.getenv('ADMIN_PASSWORD')
MAX_VOTES = 20

# Créer le dossier photos s'il n'existe pas
os.makedirs(PHOTOS_DIR, exist_ok=True)

# Initialiser la base de données
def init_db():
    try:
        if not os.path.exists(DB_PATH):
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            c.execute('''CREATE TABLE votes
                         (id INTEGER PRIMARY KEY, user_id TEXT, photo TEXT, timestamp DATETIME)''')
            conn.commit()
            conn.close()
            app.logger.info(f"Database initialized at {DB_PATH}")
    except Exception as e:
        app.logger.error(f"Database initialization error: {str(e)}")

# Récupérer ou créer l'UUID utilisateur
def get_user_id():
    user_id = request.cookies.get('user_id')
    if not user_id:
        user_id = str(uuid.uuid4())
    return user_id

# Basic Auth pour admin
def require_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization
        if not auth or not (auth.username == 'admin' and auth.password == ADMIN_PASSWORD):
            return jsonify({'error': 'Unauthorized'}), 401
        return f(*args, **kwargs)
    return decorated

# Routes principales
@app.route('/')
def index():
    response = make_response(render_template('index.html'))
    user_id = get_user_id()
    response.set_cookie('user_id', user_id, httponly=True, max_age=365*24*60*60)
    return response

@app.route('/gallery')
def gallery():
    response = make_response(render_template('gallery.html'))
    user_id = get_user_id()
    response.set_cookie('user_id', user_id, httponly=True, max_age=365*24*60*60)
    return response

@app.route('/admin')
def admin():
    auth = request.authorization
    if not auth or not (auth.username == 'admin' and auth.password == ADMIN_PASSWORD):
        return make_response('Not Authorized', 401, {'WWW-Authenticate': 'Basic realm="Admin"'})
    return render_template('admin.html')

@app.route('/results')
def results():
    response = make_response(render_template('results.html'))
    return response

@app.route('/photos/<filename>')
def serve_photo(filename):
    try:
        return send_from_directory(PHOTOS_DIR, filename)
    except FileNotFoundError:
        return jsonify({'error': 'Photo not found'}), 404

@app.route('/static/noclickright.js')
def no_click_right():
    return '''
document.addEventListener('contextmenu', (e) => e.preventDefault());
document.addEventListener('dragstart', (e) => e.preventDefault());
document.addEventListener('selectstart', (e) => e.preventDefault());
    ''', 200, {'Content-Type': 'application/javascript'}

# API endpoints
@app.route('/api/photos')
def api_photos():
    if not os.path.exists(PHOTOS_DIR):
        return jsonify([])
    photos = sorted([f for f in os.listdir(PHOTOS_DIR) if f.lower().endswith(('.jpg', '.jpeg', '.png', '.gif'))])
    return jsonify(photos)

@app.route('/api/user_votes')
def api_user_votes():
    try:
        user_id = get_user_id()
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('SELECT COUNT(*) FROM votes WHERE user_id = ?', (user_id,))
        count = c.fetchone()[0]
        
        c.execute('SELECT photo FROM votes WHERE user_id = ?', (user_id,))
        voted_photos = [row[0] for row in c.fetchall()]
        conn.close()
        
        return jsonify({'votes': count, 'voted_photos': voted_photos})
    except Exception as e:
        app.logger.error(f"Erreur /api/user_votes: {str(e)}")
        return jsonify({'votes': 0, 'voted_photos': []}), 200

@app.route('/api/vote', methods=['POST'])
def api_vote():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid JSON'}), 400
        
        photo = data.get('photo')
        if not photo:
            return jsonify({'error': 'Missing photo'}), 400
        
        user_id = get_user_id()
        
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        # Vérifier si user a déjà voté 20 fois
        c.execute('SELECT COUNT(*) FROM votes WHERE user_id = ?', (user_id,))
        if c.fetchone()[0] >= MAX_VOTES:
            conn.close()
            return jsonify({'error': 'Max votes reached'}), 400
        
        # Vérifier si user a déjà voté pour cette photo
        c.execute('SELECT COUNT(*) FROM votes WHERE user_id = ? AND photo = ?', (user_id, photo))
        if c.fetchone()[0] > 0:
            conn.close()
            return jsonify({'error': 'Already voted'}), 400
        
        # Enregistrer le vote
        c.execute('INSERT INTO votes (user_id, photo, timestamp) VALUES (?, ?, ?)',
                  (user_id, photo, datetime.now()))
        conn.commit()
        conn.close()
        
        return jsonify({'success': True})
    except Exception as e:
        app.logger.error(f"Erreur /api/vote: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/stats')
@require_auth
def api_admin_stats():
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('SELECT photo, COUNT(*) as count FROM votes GROUP BY photo ORDER BY count DESC')
        stats = [{'photo': row[0], 'votes': row[1]} for row in c.fetchall()]
        conn.close()
        return jsonify(stats)
    except Exception as e:
        app.logger.error(f"Erreur /api/admin/stats: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/count')
@require_auth
def api_admin_count():
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('SELECT COUNT(DISTINCT user_id) FROM votes')
        user_count = c.fetchone()[0]
        conn.close()
        return jsonify({'users': user_count})
    except Exception as e:
        app.logger.error(f"Erreur /api/admin/count: {str(e)}")
        return jsonify({'users': 0}), 200

@app.route('/api/admin/export')
@require_auth
def api_admin_export():
    try:
        # Récupérer toutes les photos
        all_photos = []
        if os.path.exists(PHOTOS_DIR):
            all_photos = sorted([f for f in os.listdir(PHOTOS_DIR) if f.lower().endswith(('.jpg', '.jpeg', '.png', '.gif'))])
        
        # Récupérer les votes par photo
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('SELECT photo, COUNT(*) as count FROM votes GROUP BY photo ORDER BY count DESC')
        voted_photos = {row[0]: row[1] for row in c.fetchall()}
        conn.close()
        
        # Créer liste complète : photos votées + photos sans votes
        stats = []
        for photo in all_photos:
            stats.append((photo, voted_photos.get(photo, 0)))
        
        # Trier par votes décroissants, puis par nom
        stats.sort(key=lambda x: (-x[1], x[0]))
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultats"
        
        ws['A1'] = 'Photo'
        ws['B1'] = 'Votes'
        
        header_fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        ws['A1'].fill = header_fill
        ws['B1'].fill = header_fill
        ws['A1'].font = header_font
        ws['B1'].font = header_font
        
        for idx, (photo, count) in enumerate(stats, start=2):
            ws[f'A{idx}'] = photo
            ws[f'B{idx}'] = count
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 12
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'votes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        app.logger.error(f"Erreur export: {str(e)}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=False)
